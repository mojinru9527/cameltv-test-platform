"""Report API routes."""
from __future__ import annotations

import json
import logging

from fastapi import APIRouter, BackgroundTasks, Depends, Query, Request
from sqlalchemy.orm import Session

from app.core.deps import CurrentUser, get_current_user, get_db, require_permission
from app.schemas.common import R
from app.schemas.test_report import ReportCreate, ReportDetailOut, ReportOut, TrendOut
from app.services import report_service
from app.services.audit_service import write_audit

logger = logging.getLogger("report")
router = APIRouter(prefix="/reports", tags=["报告中心"])


def _audit(req: Request, cu: CurrentUser, db: Session, action: str, target: str, detail: str = ""):
    write_audit(
        db,
        user_id=cu.user.id,
        username=cu.user.username or "",
        project_id=cu.project_id or 0,
        action=action,
        target=target,
        detail=detail,
        ip=req.client.host if req.client else "",
    )


def _run_notify_in_new_session(project_id: int, event: str, data: dict) -> None:
    """在独立 DB session 中发送通知（供 BackgroundTasks 调用）。"""
    from app.core.db import SessionLocal
    from app.services.notify_service import notify_sync

    db = SessionLocal()
    try:
        notify_sync(db, project_id, event, data)
    except Exception:
        logger.exception("Background notification failed: event=%s project=%s", event, project_id)
    finally:
        db.close()


@router.get("", response_model=R[dict])
def list_reports(
    req: Request,
    keyword: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current: CurrentUser = Depends(require_permission("report:list")),
    db: Session = Depends(get_db),
):
    items, total = report_service.list_reports(
        db,
        project_id=current.project_id or 0,
        keyword=keyword,
        page=page,
        page_size=page_size,
    )
    return R.ok({"total": total, "page": page, "page_size": page_size, "items": items})


@router.get("/trends", response_model=R[TrendOut], summary="多计划趋势与缺陷收敛")
def get_trends(
    current: CurrentUser = Depends(require_permission("report:list")),
    db: Session = Depends(get_db),
):
    """获取项目下所有报告的通过率趋势和缺陷收敛曲线数据。"""
    data = report_service.get_trends(db, current.project_id or 0)
    return R.ok(TrendOut(**data))


@router.post("", response_model=R[ReportOut])
def create_report(
    req: Request,
    body: ReportCreate,
    background_tasks: BackgroundTasks,
    current: CurrentUser = Depends(require_permission("report:create")),
    db: Session = Depends(get_db),
):
    try:
        r = report_service.create_report(db, body, current.user.id, current.project_id or 0)
        db.commit()
        _audit(req, current, db, "report:create", f"#{r['id']} {r['name']}")

        # Dispatch report_generated notification
        try:
            content = json.loads(r.get("content", "{}")) if isinstance(r.get("content"), str) else (r.get("content") or {})
        except (json.JSONDecodeError, TypeError):
            content = {}
        stats = content.get("stats", {}) if isinstance(content, dict) else {}
        total = stats.get("total", 0)
        pass_count = stats.get("pass", 0)
        pass_rate = f"{round(pass_count / total * 100, 1)}%" if total > 0 else "N/A"

        background_tasks.add_task(
            _run_notify_in_new_session,
            current.project_id or 0,
            "report_generated",
            {
                "report_name": r.get("name", ""),
                "pass_rate": pass_rate,
                "link": "",
            },
        )

        return R.ok(ReportOut(**r))
    except ValueError as e:
        from app.core.exceptions import APIException
        raise APIException(str(e))


@router.get("/{report_id}", response_model=R[ReportDetailOut])
def get_report(
    report_id: int,
    current: CurrentUser = Depends(require_permission("report:detail")),
    db: Session = Depends(get_db),
):
    r = report_service.get_report(db, report_id, current.project_id or 0)
    if not r:
        from app.core.exceptions import not_found
        raise not_found("报告")
    return R.ok(ReportDetailOut(**r))


@router.get("/{report_id}/gate", response_model=R[dict], summary="查询报告门禁评估")
def get_report_gate(
    report_id: int,
    current: CurrentUser = Depends(require_permission("report:detail")),
    db: Session = Depends(get_db),
):
    """获取指定报告的质量门禁评估结果（含 pass/fail/warn 和详情）。"""
    gate = report_service.get_report_gate(db, report_id, current.project_id or 0)
    if not gate:
        from app.core.exceptions import not_found
        raise not_found("报告")
    return R.ok(gate)


@router.delete("/{report_id}", response_model=R[dict])
def delete_report(
    req: Request,
    report_id: int,
    current: CurrentUser = Depends(require_permission("report:delete")),
    db: Session = Depends(get_db),
):
    ok = report_service.delete_report(db, report_id, current.project_id or 0)
    if not ok:
        from app.core.exceptions import not_found
        raise not_found("报告")
    db.commit()
    _audit(req, current, db, "report:delete", f"report #{report_id}")
    return R.ok({"deleted": True})


@router.get("/{report_id}/export", summary="导出报告")
def export_report(
    report_id: int,
    fmt: str = Query("excel", alias="format", pattern="^(csv|excel|pdf)$"),
    current: CurrentUser = Depends(require_permission("report:detail")),
    db: Session = Depends(get_db),
):
    """导出报告为 CSV 或 Excel（.xlsx）文件。"""
    from io import BytesIO

    from fastapi.responses import StreamingResponse

    r = report_service.get_report(db, report_id, current.project_id or 0)
    if not r:
        from app.core.exceptions import not_found
        raise not_found("报告")

    # Flatten case-level execution stats from report content
    content = r.get("content") or {}
    cases = content.get("cases", []) if isinstance(content, dict) else []

    if fmt == "csv":
        import csv

        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(["用例标题", "优先级", "域", "模块", "执行状态", "执行人", "备注"])
        for c in cases:
            writer.writerow([
                c.get("case_title", ""), c.get("priority", ""), c.get("domain", ""),
                c.get("module", ""), c.get("last_status", ""), c.get("executor_name", ""),
                c.get("notes", ""),
            ])
        output.seek(0)
        filename = f"{r.get('name', 'report')}.csv"
        return StreamingResponse(output, media_type="text/csv",
                                 headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    elif fmt == "pdf":
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        # Try multiple CJK font paths (Windows / Linux), fall back to Helvetica
        import os as _os
        _cjk_font_paths = [
            r"C:\Windows\Fonts\msyh.ttc",           # Windows
            r"C:\Windows\Fonts\simsun.ttc",         # Windows fallback
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",     # Debian/Ubuntu
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",   # Debian/Ubuntu alt
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",  # Noto
            "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",       # Noto alt
            "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc", # Fedora
        ]
        _font_loaded = False
        for _fp in _cjk_font_paths:
            if _os.path.isfile(_fp):
                try:
                    pdf.add_font("CJK", "", _fp, uni=True)
                    pdf.set_font("CJK", "", 12)
                    _font_loaded = True
                    break
                except Exception:
                    continue
        if not _font_loaded:
            pdf.set_font("Helvetica", "", 12)

        # Title
        pdf.set_font_size(16)
        pdf.cell(0, 10, r.get("name", "Test Report"), ln=True, align="C")
        pdf.ln(4)
        pdf.set_font_size(10)
        stats = r.get("stats") or {}
        info_lines = [
            f"Plan: {r.get('plan_name', '')}  |  Date: {r.get('created_at', '')}",
            f"Total: {stats.get('total',0)}  |  Pass: {stats.get('pass',0)}  |  Fail: {stats.get('fail',0)}  |  Skip: {stats.get('skip',0)}  |  Block: {stats.get('block',0)}",
        ]
        pdf.set_font_size(9)
        for line in info_lines:
            pdf.cell(0, 6, line, ln=True)
        pdf.ln(4)

        # Table header
        pdf.set_font_size(8)
        headers = ["Case", "Priority", "Domain", "Module", "Status", "Executor", "Notes"]
        widths = [45, 15, 25, 30, 18, 25, 35]
        for i, h in enumerate(headers):
            pdf.cell(widths[i], 7, h, border=1)
        pdf.ln()

        # Table rows
        for c in cases:
            row = [
                str(c.get("case_title", ""))[:30],
                str(c.get("priority", ""))[:6],
                str(c.get("domain", ""))[:15],
                str(c.get("module", ""))[:18],
                str(c.get("last_status", ""))[:10],
                str(c.get("executor_name", ""))[:15],
                str(c.get("notes", ""))[:25],
            ]
            for i, val in enumerate(row):
                pdf.cell(widths[i], 6, val, border=1)
            pdf.ln()

        output = BytesIO()
        pdf.output(output)
        output.seek(0)
        filename = f"{r.get('name', 'report')}.pdf"
        return StreamingResponse(
            output, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    else:  # excel
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "执行明细"
        ws.append(["用例标题", "优先级", "域", "模块", "执行状态", "执行人", "备注"])
        for c in cases:
            ws.append([
                c.get("case_title", ""), c.get("priority", ""), c.get("domain", ""),
                c.get("module", ""), c.get("last_status", ""), c.get("executor_name", ""),
                c.get("notes", ""),
            ])

        # Summary sheet
        ws2 = wb.create_sheet("统计概览")
        stats = r.get("stats") or {}
        ws2.append(["指标", "值"])
        ws2.append(["用例总数", stats.get("total", 0)])
        ws2.append(["通过", stats.get("pass", 0)])
        ws2.append(["失败", stats.get("fail", 0)])
        ws2.append(["跳过", stats.get("skip", 0)])
        ws2.append(["阻塞", stats.get("block", 0)])
        ws2.append(["待执行", stats.get("pending", 0)])
        ws2.append(["报告名称", r.get("name", "")])
        ws2.append(["计划名称", r.get("plan_name", "")])
        ws2.append(["创建时间", str(r.get("created_at", ""))])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"{r.get('name', 'report')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
