"""Excel import/export service for test cases."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.services.file_parser_service import parse_xlsx

# Column mapping for test case Excel export/import
COLUMNS = [
    ("case_id", "用例编号"),
    ("title", "用例标题"),
    ("domain", "域"),
    ("module", "模块"),
    ("case_type", "用例类型"),
    ("priority", "优先级"),
    ("status", "状态"),
    ("preconditions", "前置条件"),
    ("steps", "操作步骤"),
    ("expected_result", "预期结果"),
    ("api_method", "请求方法"),
    ("api_endpoint", "API 路径"),
    ("tags", "标签"),
]

FIELD_MAP = {label: field for field, label in COLUMNS}
FIELD_ORDER = [field for field, _ in COLUMNS]
HEADER_LABELS = [label for _, label in COLUMNS]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def cases_to_excel_bytes(cases: list[dict]) -> bytes:
    """Convert a list of test case dicts into an .xlsx file (bytes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header row
    for col_idx, label in enumerate(HEADER_LABELS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, case in enumerate(cases, 2):
        for col_idx, field in enumerate(FIELD_ORDER, 1):
            val = case.get(field, "")
            if isinstance(val, list):
                val = "\n".join(str(s) for s in val)
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(HEADER_LABELS) + 1):
        max_len = len(HEADER_LABELS[col_idx - 1])
        for row_idx in range(2, len(cases) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def excel_bytes_to_cases(file_bytes: bytes) -> list[dict]:
    """Parse .xlsx test case file → list of case dicts ready for creation.

    Reuses parse_xlsx from file_parser_service for header detection,
    then maps Chinese column headers to TestCase model fields.
    """
    result = parse_xlsx(file_bytes)
    raw_cases = result.get("cases") or []

    mapped: list[dict] = []
    for raw in raw_cases:
        case: dict = {}
        for label, value in raw.items():
            field = FIELD_MAP.get(label)
            if field:
                case[field] = str(value).strip() if value else ""
        if case.get("title"):  # require at least a title
            # Defaults
            case.setdefault("domain", "接口测试")
            case.setdefault("case_type", "manual")
            case.setdefault("priority", "P2")
            case.setdefault("status", "active")
            case.setdefault("source", "excel_import")
            case.setdefault("steps", "[]")
            case.setdefault("tags", "[]")
            mapped.append(case)

    return mapped
